"""
gerar_excel.py — Geração do relatório Excel formatado
Compartilhar NE — Dados Representadas
"""
import io
import base64
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import banco

_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAUAAAABuCAYAAABWQcpYAAAABmJLR0QA/wD/AP+gvaeTAAAg"
    "AElEQVR4nO2deXxcVfn/P8+5k6XpRgu0tgSaZTIzybRFTBVBwPAF2iahCEoVFRV3xRWt0la+"
    "GmVpUVAEEb+iCK78wIWtSdmkqCBiq9h2MkvSJKV7WUrb7HPveX5/zHbn3jszd9KUlnrer1de"
    "r9xzzzZz5z7nOc95znMAhUKhUCgUCoVCoVAoFAqFQqFQKBQKhUKhUCgUCoVCoVAoFAqFQqFQ"
    "KBQKhUKhUCgUCoVCoVAoFAqFQqFQHHnoSHegEL1tVeUcnzBTMwxNCp5iSE1L3fNIY5AEjZTp"
    "5XvedNPGgSPZT4VC8cbjqBGADFDvSv88MJ8O4HQAjQBOBnC8yypeAXgTEz0tmDuqyrr+SW2Q"
    "h63DCoXiDc9RIwB7Vvg/DvDPxrHKPSD8gYV2S+114a5xrFehUBwjHDEBGAwuLR0a2jehp+eJ"
    "/QDQs7J6JrhkK4CycW5KEvCA1LTlShAqFAozr7sADAQWV4HEMiZ8FIDBoNZY55q/AkDvirof"
    "M+izeYpvJyDGRPsAlgD2g1ECQgWYpgI8EcCpACY5lB0mxmp9QK6uu617ZPw/mUKheKPxugnA"
    "QKBlHmv4GhiXAShJpTOoe/aMgfp169bpXSu8J2oQMQDHJW9HiNEhBT9J0NfX3NC7p1A7vBRa"
    "r9d3JkCXAnw5gOmWLOsNQ15S993u7eP24RQKxRuSwy4A6+sXzTKg3UyEy3K1R4RLI6H2PwBA"
    "9/Jar0dodYahh2pv7HnxUNredlXlhNEJFZ8hxv8CmGa6tUdAXlK1qvvvh1K/QqF4Y3MYBeBS"
    "LdAweCUTXwvG1ALdeDLaueb8Ylvou8ZfLXVuBTCsoewPc1Zv2ueUr3tZ7Qzh0e4CodWUfECw"
    "bKpa3f3vYttVKBTHBodFAPqCF55KLO8E8FaXvdjvwWBlKLSu3012Bqh3he86AFcDSPkF7pWk"
    "nee9Ibw5V5melb5vEeNbpuQ9Uhhv9V6/ZZurfioUimMKMb7VtQlfsGU5sfwn3Ao/4BmDuNGt"
    "8AOAnuW+zwJYiYzwA4AZguWPcpUhgGtviLWB8QUAnEyeKVi72W27CoXi2MIzXhXNnXvRzLh8"
    "/m4wFrsqQDCIcV2kc+K1wP1GMW0R8EWndAZ7C5WtWR370Zblda8QiQYAfSzEX4ppW6FQHDuM"
    "iwD0BRefFWf9jwBOdFlkH8CXRjo7/jymBgVkWoczQYQ1borXru763ZjaVSgUxxSHPAX2N7Re"
    "TiyeALsUfoReKegd0dAYhR8ABl8H2ERgH8VLvuWUX6FQKJw4pEUQf0PLFwHcUkQ9XR7COaFQ"
    "++5DaRcAtqzwtRDhCmJoTHjWKJV31rV1HzjUehWKY4lgcMkpOutngzFPLzG+v2XjY3tz5fV6"
    "m6doZaIJ4LNA8tfRzWs3vn49PTKMeQocaGheycD1RRTZJhgXhDoPXfgBQO2qWDuA9vGoa+y0"
    "icDcf74dUjYxyAvQTAKXMqARYDCgA9gKRhhSeygafbj3yPZX8Uamvr51Lgv5DTCdwMCrHjK+"
    "EQo92u2UNxhseZPO6NDZeDNAAAEeo+RRADYBOG9e67RRKdeA6a1gTsgEFj0AlAB0wl/f+h0G"
    "/6/rAoSXoMsLwtG1W8fS3tFGU1OTZ/dLEz7D/PwylpiTUYA5PS/Pmp8TAM24xR9s/jsZ2lcj"
    "kUeUA7aiKBICjR8D06xUmi61GuTwttAJlWC82ZzGUjr640opJoP5jKy8wCnj0O2jnqJtgIGG"
    "ljZQEcIP0AF6XzS6NlpsW0cjPt+FJ+3aW/EsM90GYE5RhZnOkCQLrlQrjl0aG5dUBOa2vjsQ"
    "bLmimHK6xPsAzMpKJCyonb9whlN+YWTtfEpkJ+G0Rx66nuVOlsyLimL690alKA3QF2z9JjMX"
    "t9DA/I1oZ/tTRZU5Sqmvb5kjSf4FYx8deyZX7Ll3PPukeGMQDC7yGtLz9f4h430ApgBAVVXT"
    "vX1964bdlGfILrLqK4QhGtT2O+XXQaXCsk7IzOVOeT2eUU1niwxkOOY91nAtAH3Bls8R87eL"
    "rP+BaLjje0WWOSoJBhdN1xmPI4fwY2APAY8z6DViOQBBkwGUgXEmgPpELlq+YcOG+OvWacVR"
    "gwFxLRNfZk6bMKFsGoBdbsrHwmvb/Q0tPwHwmXSixLe6uzscIxtpGhNbwgELsmt6ABAXpR4y"
    "sl1xcwnLYw1XAtAfbL4UjFuLqpnwkgfGx2Exh9U1tJwuwJcCYhYgT0jkpb1grB8pLf1l3wsP"
    "vFZUO68TOms/B1DncOsgE1ZMLt/z01zCLTD3wiZI48xIZ/v9h7eXiqMVZrJNVXXNcxxcCkAA"
    "iHa2f9YbbL5NY6oTrG0Lhx/+V74Grc4ZkslRAHqMUc2wyEYhxj0u51FJQQEYCLTMY8Y9KNJe"
    "yJLaQuFHX01dexuaT9OI7gSjMfFgGOkHlBCRHyobHbm6vn7JkrwP9ggQmNtyIUtcbL/DOwVr"
    "TeHOR/IGWo1sfmQdgHWHpXOKNwaME6zOYixl0VpWd6ijE0Bn4eYEWV1lc2mAhiDNengEM43z"
    "Ntmjk7wCcP78hRNHdNwHFG0Q3Tt4sOIXqYtAsOUDzLgbnIkD6AzNlsJY653f/JbujR1HS7w+"
    "YgPXO3g6DgoWi8Lh/MLvUKivv7COmacAgKbp+6XEQDwuBrq7O8bJ37FNzJ37rxNHYEzX9MQz"
    "lpJ2x2KP7MhRgGrnLqwsNUrLpdSFEGJYiIm7QqH7Rw+1J8Hgouk6USXpWgkRHfB4+OVNm9Y4"
    "RvdxQ2Xl0gnTpg2WjwDTJEsSJFjT2Riv/mZoE755/6wiKWeRIYb1UrnH8bdL6RiXaTzC9c6p"
    "4mG2/WKZpaMAFEaJR5JtN2q6vHduc61HF8cBPBr36K+V8+i+YvbuW6mdv3CGZpTOEkbC5UaW"
    "iAFJen/3xkm7it0Wm4+qN198XPlwvBagUt1j7NX7h3ZYba55HZgDDS03M/CVolsmvj4a6rgG"
    "SAu/X6EIDZIYt0fC7Z8vut3DgK9h8XkE8YTtBtM10fCaYvwg3bcZvPBUYr4B4JYcWV4G4UEW"
    "2k2xTQ9H8tXlr29eDUHnAwCBXxweGPpAVRX03S9VvFcyrSRw0KHYNhDdbozIW7q7O0a83uYT"
    "PaX4ZDJad6Ulr07As8S4LhxufzznZ5rXWkOS70snSDADt2rAbkn4HIBW2AfkLgav1Vi7zc1A"
    "k7DTihUAvQuAF06/b4IBxhYC/ugRnls2b34ob5Bd8/eXIhqaeLq3oX++RnQVmC4BOHt1lbBf"
    "SDqromJ3tH945iYwKpD43qz9iYOwPek9dRAEPVV/QhAs1fzBgX/k7BxjZ7Sz/SKnW7765ouJ"
    "6E/Z3eJlkc4OW/APX3DxW4nF85bk+0ny7SzoagCLHfp+AAmb9w9TEd3zUV+/aJYU2ncANINx"
    "Uo5sowT8hZnu9YiKXxUaqPwNLfeBUAMAxNCZ8Fw01P7l5GzzagCXgm1a7zZjlOtSttOcAtA/"
    "t2UBGM85VFAQkpgfibRv8s9dPB9S/AMoekXphWhn+2nFtns48Adb7gXjfZbkvcYon5LLAD1W"
    "KiuXTpg4deAmAJ92+b3HwfylaLjjjlwZ/A0t9wFYakr6KxILOW5ceLaC8BckfkgTCmVm0M2x"
    "zjVfg32bIhoaFgUNaI6hylwQB/iH0c5Jy/NpCP5gy2IwOoqody+TfE8stPZvOeu0f38g4PsM"
    "fB5AaY5i4WhnexAA+xta4ijK24L6o51rJmfabz1oE7AZtkQ72x3dqnz1F15CJP+Ylch8dTTc"
    "8V1r3kDgwjNYyGctyaPI/fmyaiXgB5HO9mVweO4pEusI5NoGzqBuD/SLOzsfDeXK42to3WwZ"
    "wAcJWM/AOblrpt9HO9ekn2curYwgcctYhB8I0UikfRMAQGpfR/HCD4Ddh+lIwYwmWyLxb8db"
    "+NXOXzhj4pSBdWBcWcT3XgKiHwcaWlYU0dTZcO+/OAeMD7kRfgBA4K/665u/VERf3FIC0DJ/"
    "w+C9TU1NOYVJmab/FUAxz2UGsXjYG2xuKKYzyVlRTuHACW+JlDB4NVe+HKW357/OIudgIDSH"
    "KbBwtutJTXf6vbkRfgBADHzFX9/yjby5dM8Gl/UlKgV7DYjH6usXzSqcO01FfuEHhuAbzQmO"
    "X0gg2PJuAO8oouFMC0x/SPy3VAPwrjHVAToqXEVqas6fSsBMa7qU2VOLQ8XrbS7zGNoDAN7m"
    "cDvGoB8x6EcE/AOwn3XMwPX+YPOlY2mbgE4Qngbgxt42ysBzYF4HwPnFJPrfysqlrgSmqRM7"
    "AHqSgH8wkGdKypfu2jPx6lx3N258bIATrkoA+BUGniPgCWY8AsazIAw5FDtOYzF2Vy3CDgA9"
    "ALYh8R2ujYU7MlN98HMoLJQPInHg179BeCi7fl4Lwks5yhUXRk46rwILwzndxGtgrAdhAxLb"
    "Ox0qxzXeuc21uSqIRh/uBae31ukAugD6GwFPAPgnEhqntdLZkjzXFehbHvgVJJ7NiwAOAvTb"
    "6Ob29eYcDqNpm2B+fsyNCpEISTVv3uCUUSOn6p4XYplz1AsFq96k6aKSDaNEE7ynLvZiHzkI"
    "hfFAqyibDWnX6ksFYuPaThldA8YZluQ4EX8lEjr9x0Bb+vPV1ze/XRLdCWCuKS+B6fZ581qf"
    "dL1wQPg5dPm9SHKHTmXl0gkVUwavJfBX4Ggaod+TLP1MNPKnV1Ip/oaW94LwW4vGOr1i6tA8"
    "bIfVpuTEMJi/MmvG0J3r1q1Lv1je+c2VwqDLCVhuO06B+Jv19S2/DofbHbdVGiX6xz0jYprT"
    "zqNgsGmSzhO/BMHfzu4zL/R6m6cUubg0AvCV0VDHL5Bn6hft7HgXAPgbWnoBVJnvScJpXaH2"
    "F/I1Eg11XAXgqqQ56T+W2zkFoDSkJMp+jCycZxZSsIecrWFbQLx81olDD6SeTyBwyfGsjdwI"
    "xsctecs8kj4BIOdsxCjhVhiYUoZJ3Vb7XjC4aLoB7QfM+HB2Kb4EwCeQ5zt2YB+IroiG2h8u"
    "VM4mAP3Bf7wbTIEiGstGL40CAFF8dIxbjSVpWGVN7K6tPVn3yB8gjosB1kgISAARX01npzQ+"
    "1NC9ddxdZwTL6U6yYP/+iY7e92MhELjkeObRL9ufE30wEmq/HxaTVjjc8VwwuOidOmv/QvZU"
    "dsaoLi8DkNMemKka/y8aav+EOWn79vuHACzzNbQcT8AV5nsM7Jk8YfcHrH6O0c72+/wNLZcD"
    "WJJVv+QaoLAAZODeWLjjjmg4Oz25irraP3dxO1j8HdleCKUSdDlyBOJIRjtxjHiSXLm83h9s"
    "ngCQecrmEWXsR0ITKQzBIPDSSKjjYVf5E9g0HAa7Xo3WJBsO0i6nACQhEstMJgSz44xPCNKs"
    "TtMADhij/Jbu7o4D5pEkkhgAP+EPto6C2XqE7bk5PwDSz9WRUOjRVwFc4W9oqQTwP6Zb0/z+"
    "JVVFBBKJg+g90dAaV7vP7F8I55bgbhpPfkHYuPGxAQBbiq6BsTyyueMxc1I4UNuoa3I9GO9B"
    "dhh8ELhBCPFkqKZm3DdvsxSOU7nt2+93tX3JDZJGWhyM3C9EO9fkNBiHQo++Cs62ZQAAiJY6"
    "ZLfBjHCue5RY3bOm/T2XkzcBtoHHvRNtfgEQ3bx2I5hvs7VJ3OSufmfIgH21WlKN6woYdxYp"
    "/EAOU8cSKQ7R145za4DSSTja7YLJvDbNkIHOfBqxMSKvguW3wg6uPkXCTHjEmiiEbjND5YTo"
    "Z26FH2BR0fzB1nPB/BbXjdkax2B2X+haZr7bfQV8k3XrXDhQVUVSrgVwQp6Cx3k8/GkA+Q2x"
    "RSIENHZQoBsbGz3jtaVNgM5mq/ZHvK5QOUPgac2mNOItyHiZ52szp2GeGIb1NWFwzhhykngH"
    "WQpIyQX8PZNtETnbk8xtC+2PxDLL7scodMog4PdfNJs8+jsZ/DYwNYIwI+GHyuVMNM36DQly"
    "LwCJuOgwbMzQrZMJQxS0vRXqSe5FEMGG9TMyC0cByCwEkTUzXs7Xcnd3x4i/oWUQyX3NyTIF"
    "n0vi3KD1CyDlfBKYA2AOmKpBfBIYs8H2wdMg94uiBDzoNi9gm6Pyp4spbIOzjb2R0Jp7fA2t"
    "8wj81QIldQIvd/JREhI3c37hl2ia4C+yt4UhHHQSJfv3z5gAYFwEIBOfbG9W5LULAcDw/km9"
    "E6cMWCrDVJ/vwtl5HJlTDeQ0yksiSZYPTUS5tydK3gGLrYngTgCCcxjUTZSgv1O3++HnfNHq"
    "61vnSpLXg/RWZmiW3UZIHI/l2Bf3TsmG2Ok6b7pZ1u1b04xD0wCJc35/QsKQVnEnnDVATUhh"
    "9ZsWwsXqdcKn0kzeDROBYOsi5ud/AEZ99jDNeYdsAqYX7EsSKVDUxoT0A/D7L5oMtthyisYa"
    "UgKIda5ZBvDlcJ4O72LwT1nT5jkJv6eamjwMcnT0tEN5R6yxoEM6rkiKCZ7Z49WG4yozy5wa"
    "V4qkzc4mhKlUP95FszkFIJF9WsWcrdlnIexTZhC50wBzrSiaSNrtsvqbSwP01zd/VhJvAOii"
    "Yl24mMiF9pJA16TTanJeyMGzQRAf2pk8LHJqgAZJ1wuDTtvemB1XzC2ZbNPsnJ/HX9/6HWbu"
    "QDowiHuoCA1wcF+F673VgLnDmvEuFL/lzQI5+g5FOzt+A7T9zu9/ro5Im6QLKTGK7d3dHbmW"
    "9wEAldu3VxjC7UqKLMYB1hXdm0/v9Tc8PwBgojldSOM0AHl3YLiFwWzVDEhQ3u8lkxG6fXuh"
    "cOG/ldv2xgzDpjiAc2q7miTdqmm41QAZDqZ3Z0aBzNSIHH6n/obmDwJ0O5yd+4fB2AzCy2AI"
    "EObBGlsP7FoAarqnaAHIhGH7tFtzpynnwmGwStctPAbbZKCzBpiYAmfndbVAY9cAHQcdX7Dl"
    "c+CcMUQPELCdgd0MVBDgg0XjkxBuNcDXkoqBa9LChVmeb102HwOTvN7mMmcn4TYZjaKooKje"
    "7u7+qK96CCjoiPtPf7TPZjw9dNokuOU/IJxpTuXEzpDxOVmOMGK31RTeJN/Y2FjSP2T/XoRB"
    "hRdoiPJogLYfdcJ+lQNDsGGzAbrVAHNEJ3Eg76JKTc35UwG6DXbh95RkuvmkmQOPmt1sklvM"
    "XjHbrCiHVknEhn16OFr8IpiTFq3LQxKA7Og7l0BKtg1kuWyAwiF0FrlwKGcmw2IusT3PxHG5"
    "xmqHOe5fJfC1rs72rK1+dXNb64XkrGAPBOlSA+TcM5UcpFVfEvT2Ygs7QB4PFRclOV9lgATh"
    "+wWybWfB76UinUKL6MWTDolLAoHWxnGpnsm2vE+gNxUqNjx8gtP3zBUVosdFozlfHGK7VsYi"
    "92KFZPs9txqgdDEF9HqbT0SBXQml5aWXwbZ7iNqjne3ndYXXrMkWfkBT00sEIGvlndl5BZMd"
    "ppkej2csXgC2l1OSOCQBKPIIQCHs7wOR83NhlvYpMERBDdDBXGITgHE2rnDwclg7a8bg/1iF"
    "HwBoDgZfArnUAHMvCuVCAIlDUcDwFVvYCanxO8ejnhS7Zs1pI+CncDaTPkgoOa0+0tdnvdHl"
    "9daG66qXhb3VCzdVV7tfRreiGb93SBUs+GdF73hwgAmbrGkkqalQubgUZ9oSCbENGx4uOArK"
    "PFqiJPsLLzj3zhxN2KdhbjU7QmFNUZQ47Eii7BVKJodzMVg+ghym9ZdeKquy2QgFHG2nzNIm"
    "4A8eHCh6Cmz1kEhguN1uloM8A5lhtwEyk/Pv1cEGSCQLCnmy2wAJVi2c2fZsmOlm66CUvqex"
    "02KUG7s2UsEkikEAwKiUDcgTGKEYiKl1POpJce66dbo/1vtpAk4D4TowfkPgNsGyIRDrvdgf"
    "i9kWP8J1NR8whPEvInyPBB4tKcHuiK9me8Rf/VDYV9MW9lYvifj9k53as5I8GvAPDrfePHHK"
    "wG9dCkHyNzR/0FffeqX1hibJNnVnwvuCwUV5Rz0iusKWKOk+e06HsiK3EZ+c/MfyrBprMrcd"
    "ygV5V14bGxtLiMjul8p4JutS2rUGBuUM12RAO8teJzn2hewDguzrW1f8PnCnKbCgXFFRXFZJ"
    "OQWg4TQw5bDxOy2CUL6Fr1Q5sv1WGLZBx74bTIPMHUpLkm1mJdmlAHThVWAl8cH50B5EFoQL"
    "/f4l1eNWXxJ/rPc/gWjv/wa6ei/3x/q+7evaanPm/c/8mRMjdVU/J+LfwOyfBADgk8BYQuBv"
    "kcBDLOMXuG1bML4KhykMgIsnThl4OhBomZerbGBu65n+YPMzAP2aiH/om7cka5dNOLxmMxIR"
    "WsycoLP208bGRkcNyVff/GFYve4JQwJ64V0gAITMbcRnhxdHSi5e43EF5/ydNDYuqegfmnk3"
    "HPZHk6RbshPYNgiScJ7RzJ+/cCIzvuzQl0mNjUvs0y+7ljOA4rZlJatnmysRsVhYdD1ZdeZ2"
    "xdJkiZMjtOOgLxzcY/Ku/Gfat7bh4EFg9zmVRI4rwcHgoukgaVMSyK0GOAYzmAcAmFBp9YMc"
    "MwwNHuMXTU1N5+dScw8HUV/NWTzEd4Ecw9ZnwcATga6eB9zWHQ63b/XVN3+WiO5xuP1W1vBv"
    "f0PrOrB8hCESIZ8EzyfGu1nymcg4ynnI0G+EJUgEk1xJLJ5G9s6c9wwMzXzCF1z8DUOTMQAg"
    "XTtJAz4Aoqtsr6DkleHwo65cAOJCzzm9IcnS5tdHLhZWxgSd5g+2XGuAf+cxyvdIGS+Hh2uI"
    "cUH/kPEhWPbOJorgV5HImqzIIoLEBrZ6rDO+7m9o6Yl2tt+NpMAKBFrmjeq4E8CpTr0ZGIjP"
    "ANCXVY3dzmVxvnQHCfqP3ameL/U1NN/KQtxRTtgdj/OJLOhtAI9GO9tdaPP5NUCHd/rNDlnB"
    "zMI2c3XzOe0LZnatU2CLw3Bxk7+htd+828k/t2WBIfFTAA4uZjQrEVylYLDUoqVY8hDkwo7G"
    "RcF45669Fb8A2j5i3sh/OFg/e3bFpInlNzD4C7Adm+XIXtbkh4sNoBALd/zSH2yZC8bXbDcZ"
    "GsDngei89KpYzkdBS3zzlgTMgUxjobV/89c3fw+JII6manEOsfirR7d8LPsOkKejnaffat03"
    "nIt8bhxEMGzVu7AHjREC4xoNdA2LEaQnYrmNMbH40OgXrImjQyMPlJSXvAKQWVMoBXCXv6Hl"
    "WwB1A3wK5wqSmkSS5zRYBKDDzpgxCUDWPY/Ao++37pYg0BdI8hdGgeTwxwDobwBcCMDcNsAS"
    "xAccTrv0+hpaz7YGMGUmYXUAITEGDZDs75QH2p06jKuQ7Uo2HeD7/A0t2wmIMFAHiTm5pRdP"
    "CswdPD2yGdaYhZbmixeAyZ/cuOl/Zi73Nzz/q1zTuEOFARH1VX940qSyThB/Ce4iTo8Q6D0N"
    "4a1FOUumiIbav06gL+JQdoEQ7hs9eLDPVnf49JUA3TWG+jZ4YLy7mIHG4zFyT4Ht0xpIKQ7T"
    "FNg9BHQKNpp6ep6wBaLo6XliPzM+AbtNCgDmAHweEgdaZU6wT4SxykZI6+Z+JzvXmELBR6MP"
    "HQTjmy6zu41Kk1MA6nrZHjgMlQS+zzu/OSuqt3CIE8jShUuJVTtm2AbKUOjhF0F8BZwVjkoG"
    "zoeL+JTM/NGC/RkDAnC0c4wXH+gfnLnON6/V/UbzAoSCwdJwXc27o77qfzNwD4o4nJyAT/lj"
    "PTmj/7oh0rnmNg3GaQA/WlxJ+hszXRANtV/mfBZsm4x2rvkEEz4PQuFoMwQDhB9PKtfOSUbS"
    "cM3Q0GhugWb1iAUAYRwuDbAwBAPEt2g0eHq+KX4s3PFAMlhG3u+CgT1EWKpr+lsI+HdWU4wL"
    "/P7Ffkua9SUfkwYIANHO9tsIWI1Csw9il9u5cmuA0ehDB8k5Is9EYbDVCdymFQvN40IDzHaB"
    "YtArTtmioY7fA/QRFBbsLzPT+0lgCWDZi8x4bzLG6LiS8sMa1+jGWRDOJMmb/cHmVfGh+K1O"
    "I3ghGNAi3jlvgxAfpPjgZSDXRlFTN3CDP9b7y2LLOZEM0704EGiZx4I/TKCzOWFXMjsw7wJj"
    "ExE/LSAfzBfa2wTHQu23+/0X/ZKF/j4iLEbCbjMdADNoF4EjBP47Sfp9rph42TXyeiR2uu8D"
    "014WvKenZ1pOLYaE2Mhs3AKJEQgMsBSvQC/JuetlRPABj8T9DAwQ0RAgX5Vg+xkqzjwD5syA"
    "JGgys9k3UBzUJP2f24Onop3tDwYCl/ggRj/K4LcCmEKAhxkShP0Eet5D+l2pAWP+/IVnjxie"
    "W5ixhIh7WNIdMUscQRb8LDMmEpFOzDsgRFGRjS1wpLN9hX9uyx8gcSUIC5E4L6QfjH0g2g7w"
    "ixobd9oKstYPwh+YuZ8EvQbJO1lo6/I1RmxcwkK7OrnFVQBYK1h8P9zZ3mXJuBNM94NwgBgv"
    "AbQTeonNPctWP/hBJnoOkiUgDhIh5+8x2rnm17VzFz7tMUo+CUI9gY8DAAYbAL3EhOdKMHhP"
    "qDNx2JJ3fvNpmk43A7gAhL1gvtZqAxTgRzn7hLxthfps/wwAAsGWK5jxi0KZx4EDIL6rhEpW"
    "5zuMhgHR6fXWC2G8gxIq8nkoYkO0FQL/0RfrW3q4AqemqKk5f+rEiWVi376K4WK35Bzr5DgT"
    "5NZoZ/vhCKGvULgisQoMfnGc3AALMQVMX45L/X+QXIljgKL+qkUsaS4IPmLURQlv1mAcamyx"
    "FH872D/6ocMt/ICELepwt3FMYduApVC8vngAwBjBJu0QfdKLJD0F6/LVnAfmjvQq1PjK4SdH"
    "ygfftSC2p+g9gorDj9M2M4Xi9UQAQDIqS2F70jjBROvS/4NdhrsqFuqYOBhfcurGPWM2WisO"
    "L6Q0QMURxrz87eYQm3FB40zEYwZc78hwC4F/Z5RMuPjk7duVHe4o5jB6HygUrkivuBHhT8xw"
    "dabEITJaUqI/CwBRX+VJDIz9ACY7Boi/4Yv2fXcsTpGK1xeHXRYKxetKWgBqGHxYx8R+p83L"
    "4wrzs8kDk8BcunAcfbBfZYn313f3PVY4q+KoQAo1BVYcUdJT4FBoXT8TxqKn5UsAABSCSURB"
    "VH5AtGsy8fWI5PhMf4n+BZIL6rt7lfA7ShkRfADA/Vl/DqHAFIrXk6w118rKpRMmThnohNMm"
    "9HFCMJ8RDnc8l9zKthsFQiIVQAfoZqNkwjeDoZDrM1YVCoUCsOyf3b79/iES9EHk2WN4SBD2"
    "z5w5tB4Auuqq3opDEH4M6gTRGYFYz3Il/BQKxViwbYKObF7zLIDD453PeCoVIosJi8ZYywiD"
    "V5ePysZAtGf9OPZOoVD8l+EYQSXa2f4TELuNXOEaIk7b/xh0ftHlgUc0qQXrY30rqvv6jtwG"
    "fYVCcUyQd9+Fv6HliwBuKZTPdWMS/kikPbZ+9uyKSZPK9qHAYTcmnpGSlzV09z03Hv1QHBts"
    "qamZqnvkuRI4RwBz/LG+9xzpPqUIBpeW6rL/49Fwh6so3YojQ95TuaKd7bfW1bduIeKfOx3g"
    "XRSE3kikPQYAFRPL3oHCwk+C8QgRbvLHeq0h4xUKxEt4IZjuIwAMeyy6I0k8PnQilYicRyUo"
    "jg4KBhHtCq9ZY3j0+QzcjUMIKECS0iGShEC+k+NeBuFWEBoCXb3vUsJP4ZLybZWVh3xK33gh"
    "S4ypzHLc49cpxhc3UZSxZeNje2Od7R+VhEYQfoYxhAWXwJ/TF4wmy+0BAPexxEX9k6fPDkR7"
    "vxSI9hZ1iLpCMVpa+vqG9MhDifSUE7k++F1xhCh4MLWZrlD7CwA+WVNz/rKS8rIlAM4FuAlA"
    "oYjPI/rISAcArG9sLMHBV98KYBsDa8H00KSh0Scz+3Zt54QrFDmgUvOORynEUSMApWYIlvZD"
    "ORRHF0UJwBTJuHe/Tv4hGFxySlzqZ5AQVST5ZCbMAAAiVEiJA0T8RCpW3oING+K9VVUzq/v6"
    "bMcEKhTFwMwTzatzhjZyWM6fGQtSQhOUXwB2Lw80CmFMJ6nFPeXaxpJ+bWTAM3wGQKcIsIcF"
    "vaZJXj9ndawnVx1dbd4pnjgaJIuJL5bOfroJ62Rv3LeAJc8gFhMAwGCjs+7GbltE8lBbsLRi"
    "ZPQ0CZxILCYw8YAAeqpWxaJOe+lDbcHS8uH4OYL4YPWqrn8AwLarKifoEyadw+ATCdIjJQ6K"
    "8viT1W2Z93vrCm+DAc0LRhmTNASos3pVLGeU8RS8FNrWWt9FnPIXFjzKxK+Ulww+Obtt57iE"
    "uHtdoqAqXl/CvuqvEnA6gAMCdK8v1uMYoj7ir/aDcS0lzjw+MFw+uMIaPiwUDE4SowOXCsJC"
    "JqoCo5TAWyXTuvI4/8ppIOvyek+UQn6XwQYI34uP4rXSElwqQWcRuAYACAhLid+62b7YWV9T"
    "J3S+mAkLCVQPyGEAUQLmMCiYyseCq+sjfX2p61hd1ZuloEvBFESi3RPAeBWEEDM/VWqIe2t7"
    "ehyD2HZ5vZWGpn+QWbyZiOeAIRl4DaANGvR7fLEXcwolIHEeNEtcGe1cc3muPD0r/OsATtnD"
    "JQAd9sVBBuEhTeIrToKwZ4VvJ4DUGR+PE1DDQK05DzE+Xb069tPUdagtOKliOP4dJvo0wE6H"
    "pf8bEh+tuTH2H3Ni1wrviRrE3uTlMwB2AGgBYI0f0E/Ajxi0AcBXAD7D2gCBfuEZHvjcyT9w"
    "jti0ZUXduwm0CnA833kfGF+vWR37mVPZYlAC8Bgk6qu+noGVycuDRkm8KhjabjssKOqrvpWB"
    "1DGT2/yx3moyne0a9VVdxqDvI/OCWXmFmT5V39XzR3NiOFC1iCStTV4OI2FrdpyeMuheWTLh"
    "I067eRgQMV/VdQz6OuBiOkmyJhDdmrahROpq7gDxZ/IU2KFJ8c667u4t5tRYXU2rJL4fgPOi"
    "CvEvAtG+j+Xriq+h9WxB8spIqOP9ufL0rPQ9DcY5+eoxNTpIMFqqV3U/nVXHCt8I8ntU6B6D"
    "5pzy3ehOANi6fN40SSOPM9BYoMEDAnJx1aruv6cSLALw0GH8smZ17CPW5C0r/D8h8KcLlie+"
    "tuaGrkPyV3a1CKJ4oxH/MTLbGSd74qW2H1nE75/MQDqdGbeahV+krvprDPotcgs/ADieiO8P"
    "19Vk+d8Rk3mLYznyvKAEvkzEB1Y73Yv5qr/LoBVwI/wAGHEtO7yW4Kfyl+CTDGFknYUTCgZL"
    "JfHPkC38diIRMDgOAMSiYBAHIUhjLrAIwlYFhJ4mpuXE+DSAGwA2naHCFQxxF7dl3llOKDDW"
    "af8AGDeD6H0E/gwIH08JPwbIoJHfZgs/eowkXUyQTQR8C5kFzimSxEPdy2pnFPiomwh0NYgX"
    "gvCgw/39TPwTJr6QQVfCfHAR4X3WzL3L6z5jFn4EbCHwZyToLAZ/DDAF0GC6ZsvKutYC/cvL"
    "mGyAiqMbf2z7joiv+v8B+BAAMPhTDNxituswj36EgCnJywOlBqVPIov4a98BlquRmSE8b0jt"
    "Y0NTp8Zm7tnjGawomSeBW5PTbEHEd0Z9vqf9sdjLibpRYX6zCegG6EtMJX/dNWvW0Ek7+uZK"
    "0H0g1CXu0+fCgZOur4/sSB+rGPXNaWbgq6ZqdhPwTV2nR/eecsrON+3uq4QUawjckMpQKrLD"
    "awWivffB4YDxiL/6WjCuSV6e3eWdc1pd99Z/A4AwhucBeFOm8/TxQLTnLgB4qqnJ86bd204d"
    "Ke0vaL+SUtfAIr8AJJDpiWyuWRVrMt/mpfhmb63vQRBSL3lNT9z/FiC6HgA2fKrRMx0HzV/1"
    "sCHl6U72PgDoWVnXQozF6fqBe2tXRc0a6tM9K/wvAPxgMsMJWqn2ZWRmE1kw6F81q6ILTL+r"
    "x3tW+F4EcHI6j8SFtTd2pU/+613h28XAn5KXZTvbZlek7HmhtuAkHombB8NICcRZlasiqd/F"
    "M11t3j9oIyKCxMBMxNQGYI1T/9ygNMBjFMH8fdNlIBKoSvteMkCUmfqCgZ9l2cJYrkLqzGhw"
    "qL9/5Nxgd3dowYYN8ZO3bx/yx3qfhxhtBZCy/00D4pnRnFFm7gsTvu2P9bQHotGD565bp/u6"
    "+l6QwOdMWUqJPfOzykCYX7pXQPJMf6z3zmBPz4vnrlun10f6+og464dvaCOu/FT90d5vAngh"
    "dS2F1pT636PrWXUQZ446TbS7ZYObYxaEII0KLIJkaYBENlsq3Q/DYHm1pT/pAMKVU+JWzfqv"
    "uYQfABDoYtPliIT8ojVPzaroQwA2prtoEpi2+hg77IsllGWn1EhmfVcGG1knAxrxSdNS/1cM"
    "x1sBTE23TbzMJPwAAHVt3QcY/CtT0oL1n2oc8+KXEoDHKL6uvhfAWJe6FhKfSv0f89U0I2Nc"
    "jkudfpi6111bOwPAWalrCdqwYGf2iltvVVW5HC2bCKAvky/LFSpLAALC5ttUwto/LHlOSv0X"
    "8fsnA3h76pqA/zPb9lKw5CzbneHRbAIwEXat5qyIv+rrUV/1rWFfzR9jvqpNSJy3nOw7pxcN"
    "vN1bN8Lki8Xg28O+mj+GA1VV1rrzwUyCudDUnTLvH8PxrOakQDMdIctpATHiYasA7MvfKdSb"
    "rsJ1q7pfcs7GfzZdBs3TbjNE7NBnftF8pYuSuPlaKzeyDjwfYS0tAJlorunWwL6XpzgukAnQ"
    "383XkycfnOqUzw1qCnwMQ8Q/ZFATADDoPRH/ybMD0W07Af5yOhPjvmBPT/pHq3v0IFikNRMC"
    "PhzxVb8bSfsXAM8IMFmzDPwClDm3magk2z8PtpXWuu7uAxFf9TCSh8kzc1pYGIZxiibMv03p"
    "GPVHEA2Ze1E2nC0Ao77qt0WBuwGuBydmmwS2md7ItGpKgNEp+QNC0CMAjk+k8SWQtDhSV32D"
    "xtr36rq7R5z6Y0ZK0gTJ/BopM6W6kn8rH0UAnpkogsnpNsoPMkZMMpB4d4FuZVZrOXd7AmIX"
    "Z55f6Zb+2hOALbbFDyYcdCh+IKvnzFmLW32o6p+DnZm22MgIQJaTKX08JPYs+OmGLOGZQjJ2"
    "kOkRaiXGcQBedspbCKUBHsP4Yn0PgdGVvCwFl3wx6qsKcOKweQCAAN9kLkOsOa18TgIwLfk3"
    "2eE+GDwxXYflxbe+BCbSLhBEGXuZ8LBlRNcOwAHJlPWCGFpGAEb8NQsYWAdkaT2jSCxmZGs+"
    "jErzZUN333NSQyMBj5uSJ4BwrSGMF8KBGifXjCyEgAZQfgFIGUks8q7kcvpzEjg9MJRhNOt7"
    "ZbZPo60VmdrO/e5bRoiKSUP9ADABnuzPI+0CkMFZU17B2bFFz21bpwNkmlHQtMy/WaNqOXLA"
    "msia8hKVjFmOKQF4DEOAhKAfZFL488x0B5KLGwQ87uvqe8FcRkq2Ttu2M/AXBp6ANaR9Vnh7"
    "3pCpxPriS0cBSKbAu2TSAK0Ck0k62nhIsG6+1ksGM+0y34HMSu4uEF/cP3n6pECst8of650J"
    "on+ZOmIT6g3h3q2+WO8iZv4oALMLUYAkPxOqqTnFqU+m9omICx14kxY0MofLDS+FBmC+qURa"
    "+9y5c1bWAEBcIJAxZU2zj8uVTQpZZbrcmlqk4DIt6/MQ7AKQWGRNizXpNPhxuhxRxg9RgPaZ"
    "Ms3eunzeNDigGdI8VQYJKqiR50JNgY9x+g8O3zNpUtl3AJwAYCIosw9bStxkzU+a3Aw2jYvE"
    "n6yP9q215ssHERlsngJrztMtNglAyeZGjVeyx2ZR7abdioGEhpJ08F6Q7g/kx/3RrR0pExkB"
    "HGbeZVJzHA8CI4DR1Xd3d21te1yTv6GM5nyCVoKVAHL6GBKxdHHwe8bUwM4CsM/ruxxA2hWF"
    "pUgvCjT+dIPeu8LHqXqI2HHKmIbxPDL23dq+lfWzqm4I7zJnCbUFS2kkfqGpTNrvUMeANCuq"
    "LOx2Sxbcbz7nzKgYcdqxcRDJ6FIMSssgg+lvwlTYECNLAPzS1gbhkswVDZaX7He0ZbpBaYDH"
    "OAt27hwE4ScOt15w2oWRXGzI2JKYlj3V1FTUQMmWKbCuDTnZikCAeeRO/++Pbu2DSetixud7"
    "q6pyTonS7XgSApCYKrPvGBvNVwxoBJxqSnKc1qfwbtmy1yO1C8FIr2AS86n5ykjJBlDw4Pe0"
    "AGTCydbVzJ6r685h4DZT0lCJxFpTYUZi90iizQJHWUg27jDl1yQb92xrC6Ztt9wGMWE0/gMg"
    "YxKQROkdJLplCkxsXwQhy2LOICY6CcD0NJkg07+t2tXRp5Dl54fv9lzj95sLblnu/xhgcuVh"
    "vu9QtsUpAfhfgOHh22ExsjNTzhMACbjLdHnerF1b13R657zFnCfmO6UmXFe9LBSsehOsyKyw"
    "acb8jXty/EA5LfQEZfpHADP4nsw1N4yU0iOd9bXpqU8i9BVnBeEoHx6OA4BmUCyrFSrJOuIh"
    "5qv6BpBl9ysNBYNp1SYUrJwe9s45g00O2EMTRitAGXsVZ63M2iEhJFF+AUhmDRA4afrx/d09"
    "K/yP9qzwdfSs8D0HQU/BJJwJvDzl1GyqJaP15VnYAADv6i3djCz3qAviI/HYluV1t2xZ6buq"
    "d8T3HBhXZm7z772ros+krkZRlvV5pGW6CwCcLRT1YJt9hw9xRgCyzGiABDBYLkPGVjkTBv9j"
    "y0rf6p6Vvs9tWe5rJ+Kfm6raownN0UfRLWoK/F9AMNS3O+yr/hUBn0wm9ew+6ZT70OW8nVUv"
    "qVilxQffAyAx+jIWCiEWRnzVryGhmU2TwDQiQNPpZFjOkGEiaZpe9uc6pJ5JjCBjJst6eeNl"
    "8trSEe1dyLjXnCcMuSniqz5AwNAAMMNirOfK7duHAcC7Zcu2iK/6BaRcXRhfi9RVz2dCN4AF"
    "nHDgzkLT+09C0v1FxEt/RoIvifqqX4owHiPCbh7hiwFOu+oIzjiOOyEkDKPAodcSRNl2fz4F"
    "gKNtkQnfrr6h6zaHO8MAKgBAEwUXQfBi2UnfqBrZeTIDKQfo44noSw5P6MmS4aEPmxOGdpaP"
    "TDg+LpHyEXVYBSaN+5F5+o4DHxP1p34S5kUdAKhZ3f3YlhW+ZQR8L9nOVGJcnWzPzC5B8sKq"
    "G2JZU/hiURrgfwms4XqkhAzR9ecmD6dyIhgK9UtNnovEhnczxyEhkNKaELE9FJpAltuL4/Q3"
    "2au0BsiULQDnb3pxn9RoMZDt6wdgCifsR9Z97MNmQSs18SEAGSdawiICPkcp4ZeYzmZsdFL7"
    "EpAI/kDg1BTrRBA+mNiRQrWmuq7zdfXk3X0gBRtE+VeBCZwlwB2yvMLAA8y4oPaGWJvjQMKc"
    "+owj0sBO230L57at06tWxT5IhPcDWO/Q7jYQPl9dFltoDVSQdEvZkbwcFWRst9YvJF4GkHKZ"
    "6XPsBHN6mivNvpBJalfFvi9YLgD49zB5CiTpB3AnKH5a1Q3d/7KWLRalAf6X0BDu3RrxV98G"
    "xrv6J037VeH8W3cx8M6ov/r9AN4LRhDJxQICv8QQewl42KNnTZcBAIYmnxAG/RmgqQTKvW+W"
    "aQ/ArzDQLQ3PNuvthnBPV8TvP414dDkDH0N6ixrtAPPfAT4AEtUAT2FgX3bZLZujvspTGSXf"
    "BnAREiGVhsF4mhm313f3Phz2Vi8kgRsBzICQfUBC+Ef81adB0pdBvDBZjgAcAOh5lnyHmwg2"
    "RJpkFPADzFZANgyVlbxjQlzOYsllHI+/5r3J7ntnhUm7UBr6oLe3exfdj0KLLom+AYwbYvcC"
    "uDfUFpw0IS6P1wxDG5Vy1FvRvZPackd+HzUoOEHwiWJkcJdTJJdk8IS8x2fUrI4t62rzfqd0"
    "WJwgRgYdNbiq1d3/BrC06wveMkzAiaVClLKHeMDj2eE0rVYoChIKBks766vnHOl+jJUtNTVT"
    "u7zeos+SZkDbOO+UaewyqMJ44GtoPdtX33x3vjw9K3z/6Vnh4+Tf4/nyKg4PSgP8LyIZcmrr"
    "ke7HWMkVu68QBBjY9OK+wjnHDyFgMItCq5NZgQwOZ38UzigBqFAcBjSuWA+8tDF/Lk4vBFMB"
    "FxbF4UEJQIXiMBAK3T+KAkKNQelN10z5V4wVhwclABWKI4RgPMoCITAGmeXThUsoFAqFQqFQ"
    "KBQKhUKhUCgUCoVCoVAoFAqFQqFQKBQKhUKhUCgUCoVCoVAoFAqFQqFQKBQKhUKhUCgUCoVC"
    "oVAoFIpjiP8PswkpGSg99qcAAAAASUVORK5CYII="
)


# ── ESTILOS ──────────────────────────────────────────────────────────────────
COR_TITULO    = "1F4E79"
COR_ROXO      = "CCCCFF"
COR_LARANJA   = "FFCC99"
COR_CINZA     = "F2F2F2"
COR_BRANCO    = "FFFFFF"
COR_VERDE_F   = "E2EFDA"
COR_VERM_F    = "FFF0F0"
COR_AZUL_2    = "2E75B6"
COR_LARANJA_E = "7F3F00"

_borda = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _fnt(bold=False, color="000000", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _cel(ws, row, col, val, bold=False, color="000000", size=10,
         bg=None, align="left", fmt=None, italic=False, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _fnt(bold, color, size, italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = _fill(bg)
    c.border = _borda
    if fmt:
        c.number_format = fmt
    return c


def _merge(ws, row, c1, c2, val, bg, color="000000", bold=False, size=10,
           align="left", h=17, italic=False):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    c = ws.cell(row, c1, val)
    c.font = _fnt(bold, color, size, italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        c.fill = _fill(bg)
    c.border = _borda
    ws.row_dimensions[row].height = h
    return c


# ── ABA INDIVIDUAL DA FÁBRICA ─────────────────────────────────────────────────

def _criar_aba(wb, titulo, forn_nome, rede_nome, estados, periodo,
               prods_comprados, prods_nao_comprados, usar_familia=True):
    """Cria uma aba no workbook com os produtos comprados e não comprados."""
    safe = titulo.replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "") \
                 .replace("[", "").replace("]", "").replace(":", "")[:31]
    ws = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False

    # Cabeçalho azul — 7 colunas
    ncols = 7
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = "DADOS REPRESENTADAS"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = _fill(COR_TITULO)
    ws.row_dimensions[1].height = 36

    for i, (lb, vl) in enumerate([
        ("REDE:", rede_nome),
        ("FORNECEDOR:", forn_nome),
        ("PERÍODO:", periodo),
        ("ESTADOS:", estados),
    ], 2):
        ws.cell(i, 1).value = lb
        ws.cell(i, 1).font = _fnt(bold=True)
        ws.cell(i, 1).fill = _fill(COR_CINZA)
        ws.cell(i, 1).border = _borda
        ws.merge_cells(f"B{i}:{get_column_letter(ncols)}{i}")
        ws.cell(i, 2).value = vl
        ws.cell(i, 2).font = _fnt(bold=True, color=COR_TITULO)
        ws.cell(i, 2).fill = _fill(COR_CINZA)
        ws.cell(i, 2).border = _borda
        ws.row_dimensions[i].height = 20

    ws.row_dimensions[6].height = 4

    # Logo Compartilhar — canto superior direito (âncora E1)
    try:
        from openpyxl.drawing.image import Image as XLImage
        _logo_bytes = base64.b64decode(_LOGO_B64)
        img = XLImage(io.BytesIO(_logo_bytes))
        img.height = 72
        img.width  = 160
        img.anchor = "F2"
        ws.add_image(img)
    except Exception:
        pass

    # Cabeçalho de colunas (A–G)
    hdrs = ["CÓD. FAB", "CÓD. REDE", "PRODUTOS", "QTDE P/CX", "PESO CX", "UND.", "PREÇO"]
    bgs  = [COR_ROXO] * 7
    cors = ["000080"] * 7
    for j, (h, bg, cor) in enumerate(zip(hdrs, bgs, cors), 1):
        c = ws.cell(7, j, h)
        c.font = _fnt(bold=True, color=cor, size=9)
        c.fill = _fill(bg)
        c.border = _borda
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[7].height = 24

    linha = 8
    fam_atual = None

    # COMPRADOS
    for i, p in enumerate(prods_comprados):
        fam = p.get("familia", "")
        if usar_familia and fam != fam_atual:
            fam_atual = fam
            if fam:
                _merge(ws, linha, 1, ncols, f"  {fam}", COR_AZUL_2, "FFFFFF", True, 9, "left", 16)
                linha += 1

        bg = COR_BRANCO if i % 2 == 0 else COR_VERDE_F
        vals = [
            (p.get("codigo_fab", ""), "center"),
            (p.get("codigo_rede", ""), "center"),
            (p.get("nome", ""), "left"),
            (p.get("qtde_cx", ""), "center"),
            (p.get("peso", ""), "center"),
            (p.get("und", ""), "center"),
            (p.get("preco", ""), "right"),
        ]
        for j, (v, al) in enumerate(vals, 1):
            _cel(ws, linha, j, v, align=al, bg=bg)
        ws.row_dimensions[linha].height = 17
        linha += 1

    # Separador SEM VENDA
    n_nc = len(prods_nao_comprados)
    ws.merge_cells(f"A{linha}:{get_column_letter(ncols)}{linha}")
    c = ws.cell(linha, 1, f"PRODUTOS SEM VENDA  ({n_nc} itens — OPORTUNIDADE COMERCIAL)")
    c.font = _fnt(bold=True, color=COR_LARANJA_E, size=10)
    c.fill = _fill(COR_LARANJA)
    c.border = _borda
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[linha].height = 20
    linha += 1

    fam_atual = None
    for i, p in enumerate(prods_nao_comprados):
        fam = p.get("familia", "")
        if usar_familia and fam != fam_atual:
            fam_atual = fam
            if fam:
                _merge(ws, linha, 1, ncols, f"  {fam}", "FFE5CC", COR_LARANJA_E, True, 9, "left", 16)
                linha += 1

        bg = COR_BRANCO if i % 2 == 0 else COR_VERM_F
        vals = [
            (p.get("codigo_fab", ""), "center"),
            (p.get("codigo_rede", ""), "center"),
            (p.get("nome", ""), "left"),
            (p.get("qtde_cx", ""), "center"),
            (p.get("peso", ""), "center"),
            (p.get("und", ""), "center"),
            (p.get("preco", ""), "right"),
        ]
        for j, (v, al) in enumerate(vals, 1):
            _cel(ws, linha, j, v, align=al, bg=bg)
        ws.row_dimensions[linha].height = 17
        linha += 1

    # Larguras A–G
    for col, w in zip("ABCDEFG", [11, 16, 44, 14, 12, 10, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A8"

    # Configuração de impressão — retrato, todas as colunas em 1 página
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    from openpyxl.worksheet.page import PageMargins
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)

    n_comp = len(prods_comprados)
    n_nc   = len(prods_nao_comprados)
    valor  = sum(p.get("valor_total", 0) for p in prods_comprados)
    return n_comp, n_nc, valor


# ── ABA RESUMO ────────────────────────────────────────────────────────────────

def _criar_resumo(wb, periodo, resultados):
    ws = wb.active
    ws.title = "📊 Resumo"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "DADOS REPRESENTADAS — COMPARTILHAR NE"
    c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = _fill(COR_TITULO)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = (f"Período: {periodo}   |   Somente produtos cadastrados na rede   |   "
               "Clique no + para ver lojas   |   Clique no + da loja para ver produtos")
    c.font = Font(name="Calibri", italic=True, color="595959", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = _fill(COR_CINZA)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    hdrs = ["Fábrica", "Rede", "Estados", "Cadastrados\nna Rede",
            "Comprados", "Não\nComprados", "Positivação", "Valor (R$)", "Aba"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = _fnt(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(COR_TITULO)
        c.border = _borda
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 28

    COR_REGS = {
        "PE": ("EBF3FB", "1A5276", "D6EAF8", "EBF3FB"),
        "BA": ("FEF9E7", "7D6608", "FCF3CF", "FEF9E7"),
        "_":  ("F2F2F2", "555555", "EBEBEB", "F2F2F2"),
    }

    linha = 5
    for res in resultados:
        reg_key = "PE" if "PE" in res["rede_nome"] else ("BA" if "BA" in res["rede_nome"] else "_")
        cor_bg, cor_det, cor_l1, cor_l2 = COR_REGS[reg_key]
        porc = round(res["n_comp"] / res["n_cat"] * 100, 1) if res["n_cat"] else 0

        dados = [
            (res["fab_nome"], "left"),
            (res["rede_nome"], "left"),
            (res["estados"], "left"),
            (res["n_cat"], "center"),
            (res["n_comp"], "center"),
            (res["n_ncomp"], "center"),
            (f"{porc}%", "center"),
            (res["valor"], "right"),
            (res["titulo_aba"], "left"),
        ]
        for j, (v, al) in enumerate(dados, 1):
            c_ = _cel(ws, linha, j, v, align=al, bg=cor_bg, bold=(j == 1))
            if j == 8:
                c_.number_format = "R$ #,##0.00"
            if j == 5:
                c_.font = _fnt(bold=True, color="375623")
            if j == 6 and res["n_comp"] == 0:
                c_.font = _fnt(bold=True, color="9C0006")
        ws.row_dimensions[linha].height = 20
        linha += 1

        # Linhas de loja (outline 1)
        n_comp_rede = res["n_comp"]
        for k, loja in enumerate(res.get("lojas", [])):
            bg_l = cor_l1 if k % 2 == 0 else cor_l2
            n_comp_loja = loja.get("n_produtos", 0)
            n_nc_loja   = max(0, n_comp_rede - n_comp_loja)
            porc_loja   = round(n_comp_loja / n_comp_rede * 100, 1) if n_comp_rede else 0

            ws.merge_cells(f"A{linha}:C{linha}")
            c_ = ws.cell(linha, 1, f"    ↳ {loja['nome_faturamento']}")
            c_.font = _fnt(size=9, italic=True, color=cor_det)
            c_.border = _borda
            c_.alignment = Alignment(horizontal="left", vertical="center")
            c_.fill = _fill(bg_l)

            for j_blank in range(2, 10):
                ws.cell(linha, j_blank).fill = _fill(bg_l)
                ws.cell(linha, j_blank).border = _borda

            ws.cell(linha, 4).value = loja.get("estado", "")
            ws.cell(linha, 4).font = _fnt(size=9, bold=True, color=cor_det)
            ws.cell(linha, 4).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(linha, 5).value = n_comp_loja
            ws.cell(linha, 5).font = _fnt(size=9, color="375623")
            ws.cell(linha, 5).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(linha, 6).value = n_nc_loja
            ws.cell(linha, 6).font = _fnt(size=9, color="9C0006" if n_nc_loja > 0 else cor_det)
            ws.cell(linha, 6).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(linha, 7).value = f"{porc_loja}%"
            ws.cell(linha, 7).font = _fnt(size=9, color=cor_det)
            ws.cell(linha, 7).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(linha, 8).value = loja.get("valor_total", 0)
            ws.cell(linha, 8).number_format = "R$ #,##0.00"
            ws.cell(linha, 8).font = _fnt(size=9, color=cor_det)
            ws.cell(linha, 8).alignment = Alignment(horizontal="right", vertical="center")

            ws.row_dimensions[linha].height = 15
            ws.row_dimensions[linha].outline_level = 1
            ws.row_dimensions[linha].hidden = True
            linha += 1

            # Linhas de produto por loja (outline 2)
            prods_loja    = res.get("loja_prods", {}).get(loja.get("loja_id"), {})
            prods_cat     = res.get("prods_cat", [])
            ids_comp_rede = res.get("ids_comp_rede", set())
            for p in prods_cat:
                comprou      = p["id"] in prods_loja
                rede_comprou = p["id"] in ids_comp_rede
                if not comprou and not rede_comprou:
                    continue
                vd     = prods_loja.get(p["id"], {})
                bg_p   = "E8F8F5" if comprou else "FDEDEC"
                cor_p  = "1A5276" if comprou else "922B21"
                status = "✓" if comprou else "✗"
                label  = f"               {status}  [{p.get('codigo_fab','')}]  {p.get('nome','')}"

                ws.merge_cells(f"A{linha}:D{linha}")
                c_ = ws.cell(linha, 1, label)
                c_.font      = _fnt(size=8, italic=True, color=cor_p)
                c_.fill      = _fill(bg_p)
                c_.border    = _borda
                c_.alignment = Alignment(horizontal="left", vertical="center")

                for j_b in range(2, 10):
                    ws.cell(linha, j_b).fill   = _fill(bg_p)
                    ws.cell(linha, j_b).border = _borda

                if comprou:
                    c5 = ws.cell(linha, 5, vd.get("qtd", 0))
                    c5.font          = _fnt(size=8, color=cor_p)
                    c5.alignment     = Alignment(horizontal="center", vertical="center")
                    c5.number_format = "#,##0.0"
                else:
                    c6 = ws.cell(linha, 6, "✗")
                    c6.font      = _fnt(size=8, bold=True, color=cor_p)
                    c6.alignment = Alignment(horizontal="center", vertical="center")

                if comprou:
                    c8 = ws.cell(linha, 8, vd.get("valor", 0))
                    c8.font          = _fnt(size=8, color=cor_p)
                    c8.alignment     = Alignment(horizontal="right", vertical="center")
                    c8.number_format = "R$ #,##0.00"

                ws.row_dimensions[linha].height        = 13
                ws.row_dimensions[linha].outline_level = 2
                ws.row_dimensions[linha].hidden        = True
                linha += 1

    ws.auto_filter.ref = "A4:I4"
    for col, w in zip("ABCDEFGHI", [22, 26, 16, 12, 11, 11, 11, 18, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


# ── FUNÇÃO PRINCIPAL ─────────────────────────────────────────────────────────

def gerar_relatorio(fab_nome, redes_selecionadas, data_inicio, data_fim):
    fab = banco.obter_fabrica_por_nome(fab_nome)
    if not fab:
        return None, "Fábrica não encontrada"

    periodo = ""
    if data_inicio and data_fim:
        def fmt(d):
            parts = d.split("-")
            return f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else d
        periodo = f"{fmt(data_inicio)} a {fmt(data_fim)}"
    elif data_inicio:
        periodo = f"A partir de {data_inicio}"
    elif data_fim:
        periodo = f"Até {data_fim}"
    else:
        periodo = "Todo o período"

    wb = openpyxl.Workbook()
    resultados = []

    for rede_nome in redes_selecionadas:
        rede = banco.obter_rede_por_nome(rede_nome)
        if not rede:
            continue

        prods_cat = banco.produtos_com_codigos(fab["id"], rede["id"])
        if not prods_cat:
            continue

        vendas      = banco.vendas_por_produto(fab["id"], rede["id"], data_inicio, data_fim)
        lojas       = banco.vendas_por_loja(fab["id"], rede["id"], data_inicio, data_fim)
        loja_prods  = banco.vendas_por_loja_produto(fab["id"], rede["id"], data_inicio, data_fim)
        precos_fat  = banco.ultimos_precos(fab["id"], rede["id"])
        tab_prec    = banco.precos_tabela(rede["id"], fab["id"])

        def _fmt_preco(v):
            return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        comprados = []
        nao_comprados = []
        for p in prods_cat:
            vd = vendas.get(p["id"])
            if p["id"] in tab_prec:
                tp        = tab_prec[p["id"]]
                preco_str = _fmt_preco(tp["preco"]) if tp["preco"] else ""
                und_final = tp["unidade"] if tp["unidade"] else p.get("und", "")
            else:
                preco_fat = precos_fat.get(p["id"])
                preco_str = _fmt_preco(preco_fat) if preco_fat else ""
                und_final = p.get("und", "")
            p_com_preco = {**p, "preco": preco_str, "und": und_final}
            if vd and vd["valor_total"] > 0:
                comprados.append({**p_com_preco, **vd})
            else:
                nao_comprados.append(p_com_preco)

        estados    = rede["estados"].replace(",", " / ")
        titulo_aba = f"{fab_nome[:12]} {rede_nome.split()[-1]}"
        n_comp, n_nc, valor = _criar_aba(
            wb, titulo_aba, fab_nome, rede_nome, estados, periodo,
            comprados, nao_comprados, usar_familia=True
        )

        ids_comp_rede = {p["id"] for p in comprados}
        resultados.append({
            "fab_nome":      fab_nome,
            "rede_nome":     rede_nome,
            "estados":       estados,
            "n_cat":         len(prods_cat),
            "n_comp":        n_comp,
            "n_ncomp":       n_nc,
            "valor":         valor,
            "titulo_aba":    titulo_aba,
            "lojas":         lojas,
            "loja_prods":    loja_prods,
            "prods_cat":     prods_cat,
            "ids_comp_rede": ids_comp_rede,
        })

    if not resultados:
        return None, "Nenhum produto com código cadastrado para as redes selecionadas"

    _criar_resumo(wb, periodo, resultados)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), None
